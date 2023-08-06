import os
import sys
import csv
import openpyxl
from . rawInputLogFile  import rawInputLogFile
from . binarySectorFile import binarySectorFile
from . csvWriter        import csvWriter


def processInputLogFile(inputLogFile):
    retStatus = 0
    # From raw input log file, create sector bin files and process them.
    if (0 == rawInputLogFile(inputLogFile).createBinSectorFiles()):
        for i in range(0, 10, 1):
            process_id = i
            myCSVfileWriter = csvWriter('.\\csvWriter'+str('{0:02}'.format(i))+'.csv', process_id)
            binarySectorFile('.\\binarySectorFile'+str('{0:02}'.format(i))+'.bin').process(48, myCSVfileWriter, process_id)
        for i in range(10, 12, 1):
            process_id = i
            myCSVfileWriter = csvWriter('.\\csvWriter'+str('{0:02}'.format(i))+'.csv', process_id)
            binarySectorFile('.\\binarySectorFile'+str('{0:02}'.format(i))+'.bin').process(564, myCSVfileWriter, process_id)
    else:
        retStatus = 1

    return retStatus

def genOutputFile(outputFileName):

    outputLogFiles = [
        'CriticalFaultDataSection',
        'FlightFaultDataSection',
        'GroundFaultDataSection',
        'GroundTestRecordSection',
        'ConfigAndStatDataSection',
        'MaintHistoryDataSection']

    listOfcsvFiles = [file for file in os.listdir() if (file.startswith("csvWriter") and file.endswith(".csv"))]
    if ((len(listOfcsvFiles) == 12) and (len(outputLogFiles) == 6)):
        for i in range(0, len(outputLogFiles), 1):
            try:
                outputFile = open(outputLogFiles[i]+'.csv', 'a')
            except FileNotFoundError:
                print('FileNotFoundError: invalid output file name.')
                return 1
            except PermissionError:
                print('PermissionError: to create output csv file named \'{0}\', need permission on current directory.'.format(outputLogFile))
                return 1

            for file in listOfcsvFiles[(2*i):((2*i)+2):1]:
                with open(file, 'r') as inputFile:
                    while True:
                        aLine = inputFile.readline()
                        if aLine == '':
                            break
                        else:
                            outputFile.write(aLine)
            outputFile.close()

    wb = openpyxl.Workbook()
    del wb['Sheet']
    for i in range(0, len(outputLogFiles), 1):
        wb.create_sheet()
        wbSheet = wb['Sheet']
        wbSheet.title = outputLogFiles[i]

        if (outputLogFiles[i] == 'GroundTestRecordSection'):
            wbSheet.cell(1, 1).value = 'TEST ID'
            wbSheet.cell(1, 2).value = 'TEST RESULT'
            wbSheet.cell(1, 3).value = 'DATE OF OCCURRENCE (DD/MM/YYYY)'
            wbSheet.cell(1, 4).value = 'TIME OF OCCURRENCE (HH:MM)'
            wbSheet.cell(1, 5).value = 'ACT 1ST'
            wbSheet.cell(1, 6).value = 'ACT 2ND'
            wbSheet.cell(1, 7).value = 'ACT 3RD'
            wbSheet.cell(1, 8).value = 'ACT 4TH'
            wbSheet.cell(1, 9).value = 'ACT 5TH'
            wbSheet.cell(1,10).value = 'ACT 6TH'
            wbSheet.cell(1,11).value = 'ACT 7TH'
            wbSheet.cell(1,12).value = 'APPLICATION DATA'
        else:
            wbSheet.cell(1, 1).value = 'FSID'
            wbSheet.cell(1, 2).value = 'FAULT NAME'
            wbSheet.cell(1, 3).value = 'DATE OF OCCURRENCE (DD/MM/YYYY)'
            wbSheet.cell(1, 4).value = 'TIME OF OCCURRENCE (HH:MM)'
            wbSheet.cell(1, 5).value = 'FLIGHT PHASE'
            wbSheet.cell(1, 6).value = 'ACT 1ST'
            wbSheet.cell(1, 7).value = 'ACT 2ND'
            wbSheet.cell(1, 8).value = 'ACT 3RD'
            wbSheet.cell(1, 9).value = 'ACT 4TH'
            wbSheet.cell(1,10).value = 'ACT 5TH'
            wbSheet.cell(1,11).value = 'ACT 6TH'
            wbSheet.cell(1,12).value = 'ACT 7TH'
            wbSheet.cell(1,13).value = 'BITE DATA'

        csvfile = open(outputLogFiles[i]+'.csv', 'r', newline='')
        csvReader = csv.reader(csvfile, delimiter=',', quotechar='"')

        i = 1
        for row in csvReader:
            i+= 1
            for j in range(0,len(row),1):
                wbSheet.cell(i, (j+1)).value = row[j]

        csvfile.close()

    wb.save(outputFileName)

    return 0

def myMain(inputLogFile):
    # Delete intermediate sector bin files in the current directory
    listOfbinFiles = [file for file in os.listdir() if (file.startswith("binarySectorFile") and file.endswith(".bin"))]
    for file in listOfbinFiles: os.remove(file)
    listOfcsvFiles = [file for file in os.listdir() if (file.endswith(".csv"))]
    for file in listOfcsvFiles: os.remove(file)
    #os.remove('.\\rawInputFile.txt')

    if (0 == processInputLogFile(inputLogFile)):
        if (0 == genOutputFile('{0}.xlsx'.format(inputLogFile.split('.')[0]))):
            print('-- main function sucessful')

    # Delete intermediate sector bin files in the current directory
    listOfbinFiles = [file for file in os.listdir() if (file.startswith("binarySectorFile") and file.endswith(".bin"))]
    for file in listOfbinFiles: os.remove(file)
    listOfcsvFiles = [file for file in os.listdir() if file.endswith(".csv")]
    for file in listOfcsvFiles: os.remove(file)
    #os.remove('.\\rawInputFile.txt')
