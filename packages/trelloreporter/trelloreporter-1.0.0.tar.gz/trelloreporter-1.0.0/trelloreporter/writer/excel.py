import asyncio
import logging

import uvloop
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles.cell_style import CellStyle

from trelloreporter.rest import trello

class TrelloReader(object):

    def __init__(self):
        """

        """
        self._reader = trello.TrelloRest()
        self._cards = dict()
        self._lists = dict()
        self._members = dict()
        self._boards = dict()

        asyncio.set_event_loop_policy(uvloop.EventLoopPolicy())
        self._loop = asyncio.get_event_loop()

        self._logger = logging.getLogger('trelloreporter')

    def configure_api_parameters(self, api_token, api_key, api_url, api_organisation):
        """
        A function to initialise the API parameters of the trello reader object
        :param api_token: the API token
        :param api_key: the API key
        :param api_url: the API URL
        :param api_organisation: the API organisation
        :return: <none>
        """
        self._reader.configure_api_parameters(api_token, api_key, api_url, api_organisation)

    def populate_boards(self):
        """
        Boards are the highest level concept within the Trello workflow.
        The Boards API allows you to list, view, create, and edit Boards.
        Each Board has a name, description, a set of members attached, and an ordered array of Lists.

        Populates the boards found under the api_key's endpoint as a JSON dict.
        :return: <null> initialises the cards to the json list of boards
        """
        self._loop.run_until_complete(self._reader.populate_boards())
        self._boards = self._reader.boards

    def populate_cards(self):
        """
        Extracts the cards found at the api_key's endpoint, storing each
        card as a JSON item in a dict
        :return:
        """
        if self._boards == {}:
            self.populate_boards()
        self._loop.run_until_complete(self._reader.populate_cards())
        self._cards = self._reader.cards

    def populate_lists(self):
        """
        Populates the lists found under the organisation and stores them in the lists() dictionary
        :return: <dict> of Trello lists
        """
        if self._boards == {}:
            self.populate_boards()
        self._loop.run_until_complete(self._reader.populate_lists())
        self._lists = self._reader.lists

    def populate_members(self):
        """
        Method to populate the members of the Trello organisation
        :return: <null> : Initialises the members() dict
        """
        self._loop.run_until_complete(self._reader.populate_members())
        self._members = self._reader.members

    def get_members(self):
        """

        :return: <dict> : return the members of the organisation
        """
        return self._members

    def get_list_cards(self, list_id):
        """
        Returns a dict of cards based on a filter argument "lane"

        :param lane: <str> The lane in which the cards reside
        :return: <list> The cards which all belong to the lane
        """
        list_found = 0   #Marker for list found: this remains zero if there is no list which matches the search
        my_cards = dict() #A place to store cards belonging to the list
        for board_id in self._boards.keys():
            for card_id in self._cards[board_id]:
                if list_id == self._cards[board_id][card_id]['idList']:
                    my_cards[card_id] = self._cards[board_id][card_id]
                    list_found = 1   #Set the marker to one

        if list_found == 1:
            return my_cards
        else:
            return "List " + list_id + " not found"

    def get_boards(self):
        """

        :return: <list> : returns a json array of boards
        """
        return self._boards

    def get_board(self, board_id):
        """
        :param: <str> : the ID of the board to return
        :return: <dict> : returns a json array of boards
        """
        if board_id in self._boards.keys():
            return self._boards[board_id]
        else:
            return "Board " + board_id + " not found"

    def get_board_cards(self, board_id):
        """

        :param board_id: The boardID of the cards to return
        :return: <dict> All the cards which belong ot the board at hand
        """
        if board_id in self._boards.keys():
            return self._cards[board_id]
        else:
            return "Board " + board_id + " not found"

    def get_board_lists(self, board_id):
        """

        :param board_id: the boardID of the lists
        :return: <dict> : returns a dict of all the lists in a given board
        """
        if board_id in self._boards.keys():
            return self._lists[board_id]
        else:
            return "Board " + board_id + " not found"

    def get_lists(self):
        """

        :return: <dict> : returns a dict of lists
        """
        return self._lists

    def get_list(self, list_id):
        """
        :param list_id:  <str> : the ID of the list
        :return: <dict> : returns a json representation of a list item
        """
        for board in self._boards.keys():
            board_id = board
            for list in self._lists:
                if list_id in self._lists[board_id].keys():
                    return self._lists[board_id][list_id]

        return "List " + list_id + " not found" #At the end of the loop, there is no list found

    def get_cards(self):
        """

        :return: <dict> : returns a json dict of cards
        """
        return self._cards

    def get_card(self, card_id):
        """
        :param card_id: <str> : the ID of the card to retrieve
        :return: <dict> : returns a json representation of the card
        """
        for board in self._boards.keys():
            board_id = board
            for card in self._cards:
                if card_id in self._cards[board_id].keys():
                    return self._cards[board_id][card_id]

        return "Card " + card_id + " not found" #At the end of the loop there is no card found

    def get_card_summary(self, card_id):
        """
        :parameter : str : The ID of the card to return a summary of
        :return: <json> A summarized version of the card object in cards[board_id][boards]
        """
        for board in self._boards.keys():
            board_id = board
            for card in self._cards:
                if card_id in self._cards[board_id].keys():
                    self._logger.debug('Summarizing card ID: {0} {1}'.format(
                        card_id, self._cards[board_id][card_id]['name']))
                    my_card = {'Board':'',
                               'Description':'',
                               'Status':'',
                               'DueDate':'',
                               'Members':'',
                               'LastUpdate':'',
                               'URL':''}
                    my_card['Board'] = self._boards[board_id]['name']
                    my_card['Description'] = self._cards[board_id][card_id]['name']
                    listID = self._cards[board_id][card_id]['idList']
                    my_card['Status'] = self._lists[board_id][listID]['name']
                    my_card['DueDate'] = self._cards[board_id][card_id]['due']
                    my_card['LastUpdate'] = self._cards[board_id][card_id]['dateLastActivity']
                    my_card['URL'] = self._cards[board_id][card_id]['url']
                    my_card['Members'] = dict()
                    counter = 1
                    for member_id in self._cards[board_id][card_id]['idMembers']:
                        try:
                            member_user_name = self._members[member_id]['username']
                            my_card['Members'][member_user_name] = self._members[member_id]['fullName']
                        except KeyError:
                            my_card['Members']['Removed Member ID {0}:'.format(counter)] = member_id
                            counter = counter + 1
                    return my_card

        return "Card " + card_id + " not found"

    def get_board_summary(self, board_id):
        """
        :parameter : str : The ID of the card to return a summary of
        :return: <json> A summarized version of the card object in cards[board_id][boards]
        """
        card_id = ""
        for card_id in self._cards[board_id].keys():
            self._logger.debug('Summarizing card ID: {0} {1}'.format(
                card_id, self._cards[board_id][card_id]['name']))
            my_card = {'Board':'',
                       'Description':'',
                       'Status':'',
                       'DueDate':'',
                       'Members':'',
                       'LastUpdate':'',
                       'URL':''}
            my_card['Board'] = self._boards[board_id]['name']
            my_card['Description'] = self._cards[board_id][card_id]['name']
            listID = self._cards[board_id][card_id]['idList']
            my_card['Status'] = self._lists[board_id][listID]['name']
            my_card['DueDate'] = self._cards[board_id][card_id]['due']
            my_card['LastUpdate'] = self._cards[board_id][card_id]['dateLastActivity']
            my_card['URL'] = self._cards[board_id][card_id]['url']
            my_card['Members'] = dict()
            counter = 1
            for member_id in self._cards[board_id][card_id]['idMembers']:
                try:
                    member_user_name = self._members[member_id]['username']
                    my_card['Members'][member_user_name] = self._members[member_id]['fullName']
                except KeyError:
                    my_card['Members']['Removed Member ID {0}:'.format(counter)] = member_id
                    counter = counter + 1
            return my_card

        return "Card " + card_id + " not found"
    def write_to_excel(self, filename):
        """
         Create named styles, using the bellow color schemes in Excel
         Neutral, Bad, Input, WarningText, Good
        """
        self._logger.info('Creating spreadsheet')
        wb = Workbook()
        ws = wb.active  # get the active workbook

        for board in self.get_boards().keys():
            # Name the worksheet using the board name
            board_object = self.get_board(board_id=board)
            board_name = board_object['name'].replace('[', '(')  # Remove invalid characters

            # Ensure the sheet name is < 31 characters: Outlook fails to open anything else
            trunc_board_name = (board_name[:31]) if len(board_name) > 31  else board_name

            ws.title = trunc_board_name.replace(']', ')')
            self._logger.debug('Create worksheet :{0}'.format(ws.title))

            # Set the column headings
            ws['A1'] = "Task Name"
            ws['B1'] = "Status"
            ws['C1'] = "Due Date"
            ws['D1'] = "Members"
            ws['E1'] = "URL"
            row = 2
            row_counter = 0  # N rows
            for card in self.get_board_cards(board):
                card_object = self.get_card_summary(card_id=card)
                ws['A{0}'.format(row)] = card_object['Description']
                ws['B{0}'.format(row)] = card_object['Status']
                ws['C{0}'.format(row)] = card_object['DueDate']

                if len(card_object['Members']) == 0:
                    ws['D{0}'.format(row)] = "None"
                else:
                    members = ""
                    for member in card_object['Members'].values():
                        members = member + '\n' + members
                    ws['D{0}'.format(row)] = members

                ws['E{0}'.format(row)] = card_object['URL']

                # Apply Styling for column B data i.e. Status
                if card_object['Status'] == 'Backlog':
                    cell = ws['C{0}'.format(row)]   #Column C: Due Date
                    cell.style = 'Accent4'
                elif card_object['Status'] == 'In Progress':
                    cell = ws['C{0}'.format(row)]
                    cell.style = 'Accent1'
                elif card_object['Status'] == 'Impeded':
                    cell = ws['C{0}'.format(row)]
                    cell.style = 'Bad'
                elif card_object['Status'] == 'Stopped':
                    cell = ws['C{0}'.format(row)]
                    cell.style = 'Accent2'
                elif card_object['Status'] == 'Completed':
                    cell = ws['C{0}'.format(row)]
                    cell.style = 'Good'

                # Apply styling to column E - URL
                ws['E{0}'.format(row)].style = 'Hyperlink'

                    # Apply colouring to the tab
                if 'Platform' in board_name:
                    ws.sheet_properties.tabColor = "976360"
                elif 'Client' in board_name:
                    ws.sheet_properties.tabColor = "6878CC"
                elif 'Core' in board_name:
                    ws.sheet_properties.tabColor = "FFA500"
                else:
                    ws.sheet_properties.tabColor = "500FFA"

                row_counter += row_counter

            # Sort by list name. Order:
            # * Backlog, In Progress, Impeded, Stopped, Incidents(?), Completed
            #ws.auto_filter.ref = 'A1:E{0}'.format(row_counter)
            #ws.auto_filter.add_filter_column(0, ["Backlog", "In Progress",
            #                                     "Impeded", "Stopped",
            #                                     "Completed"])
            #ws.auto_filter.add_sort_condition('B2:B{0}'.format(row_counter))  # Status column

            wb.create_sheet("New Worksheet")  # Generate a new worksheet
            ws = wb["New Worksheet"]

        try:
            wb.save(filename)
            self._logger.info('Spreadsheet {0} successfully created'.format(filename))
        except Exception:
            self._logger.error('Unable to save {0}'.format(filename))