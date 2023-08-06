from openpyxl import Workbook
from .base import Writer


class ExcelWriter(Writer):

    def save(self):
        book = Workbook()
        sheet = book.get_active_sheet()

        for j, column_name in enumerate(self._columns):
            sheet.cell(row=1, column=j+1).value = column_name

        for i, each in enumerate(self._data):
            for j, value in enumerate(each):
                sheet.cell(row=i+2, column=j+1).value = value

        book.save(filename=self._path)
