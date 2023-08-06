from __future__ import print_function, unicode_literals
import ctxlogger
import openpyxl
from io import BytesIO

from .. import ParseException, string_field

class excel_batch(object):

    columns = ()
    header = False

    def __init__(self, columns, header = False):
        self.columns = columns
        self.header = header

    def parse(self, data):

        wb = openpyxl.reader.excel.load_workbook(
            BytesIO(data), read_only=True
        )

        parsed = []
        ws = wb.active

        for rowidx, row in enumerate(ws.rows, 1):
            if rowidx == 1 and self.header: # ignore
                continue

            for c in row:
                if c.value != None:
                    break
            else:
                continue # skip empty line

            with ctxlogger.context('row', rowidx):
                if len(self.columns) > len(row):
                    ctxlogger.exception(
                        ParseException,
                        'Expecting {} columns, got {}'.format(
                                                    len(self.columns), len(row))
                    )
                colnames, colfields = zip(*self.columns)

                # TODO: decide if this wanted/useful
                # parsed_row = { '_line': rowidx }
                parsed_row = {}
                colidx = 1
                for cell, name, field in zip(row, colnames, colfields):
                    with ctxlogger.context('column', name):
                        parsed_row[ name ] = field.parse(cell.value)
                parsed.append(parsed_row)

        return parsed

    def unparse(self, data, properties = {}):
        """
        properties - dict matching the values the excel workbook
            properties attribute can take
        """
        wb = openpyxl.Workbook()
        for key, value in properties.items():
            setattr( wb.properties, key, value )
        ws = wb.active

        rowstart = 1
        if self.header:
            for colidx, (colname, colfield) in enumerate(self.columns, 1):
                cell = ws.cell(row = rowstart, column = colidx)
                cell.value = colname
                try:
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal = colfield.flags['justify']
                    )
                except KeyError:
                    pass
            rowstart += 1

        for rowidx, row in enumerate(data, rowstart):
            with ctxlogger.context('row', rowidx):
                if len(self.columns) != len(row):
                    ctxlogger.exception(
                        ParseException,
                        'Expecting {} columns, got {}'.format(
                            len(self.columns), len(row)
                        )
                    )

                for colidx, (colname, colfield) in enumerate(self.columns, 1):
                    with ctxlogger.context('column', colname):
                        try:
                            value = row[colname]
                        except TypeError: # row is a list, not a dict
                            value = row[colidx - 1]

                        cell = ws.cell(row = rowidx, column = colidx)
                        cell.value = colfield.unparse(value)
                        try:
                            cell.alignment = openpyxl.styles.Alignment(
                                horizontal=colfield.flags['justify']
                            )
                        except KeyError:
                            pass

        return openpyxl.writer.excel.save_virtual_workbook(wb)
