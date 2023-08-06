#!/usr/bin/env python
# -*- coding: utf-8  -*-

import os.path
import re

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

from datalet.data import *
from datalet.storage.bin_file_storage import BinFileStorage
from datalet.storage.exceptions import (StorageExistsError
	, StorageNotFoundError
	, UnmatchExtensionError
	, ArgumentsAbsenceError)


class Excel2007Storage(BinFileStorage):
	"""
	Excel 2007 format Storage
	"""


	def __init__(self, filepath = None, sheetIndex = -1, sheetName = None):
		super().__init__(filepath)
		self.sheetIndex = sheetIndex
		self.sheetName = sheetName


	def __get_worksheet(self, workbook_toread):
		ws = None
		if not self.sheetName is None:
			ws = workbook_toread.get_sheet_by_name(self.sheetName)
		else:
			if not self.sheetIndex is None:
				ws = workbook_toread.get_sheet_by_name(workbook_toread.get_sheet_names()[self.sheetIndex])
			else:
				raise ArgumentsAbsenceError("The targetSheet arguments must be specified one:  sheetIndex=%s, sheetName=%s" % \
					(self.sheetIndex, self.sheetName))
		return ws


	def create(self, force = False):
		if self.exists():
			if force == False:
				raise StorageExistsError(self.filepath)
			else:
				self.remove(force = True)

		wb = Workbook()
		createSheetName = self.sheetName if self.sheetName is not None else ("SHEET_" + str(self.sheetIndex))
		wb.create_sheet(title =createSheetName)
		wb.save(self.filepath)


	def clear(self, force = False):
		wb = load_workbook(filename = self.filepath)
		ws = self.__get_worksheet(wb)
		wb.remove(ws)
		createSheetName = self.sheetName if self.sheetName is not None else ("SHEET_" + str(self.sheetIndex))
		wb.create_sheet(title =createSheetName)
		wb.save(self.filepath)


	def read(self, limit =  -1, data_only = True, read_only = True):
		dt = DataTable()
		'''
		read_only: set True when read large file.
		data_only: set True when need calculate the formatular.
		'''
		wb = load_workbook(filename = self.filepath, data_only = data_only, read_only = read_only)
		ws = self.__get_worksheet(wb)

		for row in ws.rows:
			dr = DataRow()
			for cell in row:
				dr.append(cell.value)
			dt.append(dr)

		# get cell data
		#for rowNum in range(1, ws.max_row + 1):
		#	if limit > -1 and rowNum > limit:
		#		break

		#	rowdata = []
		#	for colNum in range(1, ws.max_column + 1):
		#		cellval = ws.cell(row = rowNum, column= colNum).value
		#		rowdata.append(cellval)
		#	tabledata.append(rowdata)

		return dt

	def write(self, data, overwrite = False, write_only = True, write_header = True):
		'''
		use write_only mode to write large data to file.
		'''
		ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
		# if overwrite == True:
			# self.clear(force = True)
		if write_only == True:
			# data = [] if data is None else data
			# if self.exists():
				# old_data = self.read()
				# old_data.extend(data)
				# data = old_data
			#self.remove(force = True)
			wb = Workbook(write_only = True)
			ws = wb.create_sheet(title = self.sheetName)
			if write_header == True:
				ws.append(data.column_names)
			for row in data:
				new_row = []
				for data in row:
					new_data = str(data) if data is not None else ''
					if ILLEGAL_CHARACTERS_RE.match(new_data) is not None:
						new_data = "*"
					new_row.append(new_data)
				try:
					ws.append(list(map(str, new_row)))
				except Exception as e:
					print(row)
					print(e)
					raise e
			wb.save(self.filepath)
		else:
			wb = load_workbook(filename = self.filepath)
			ws = self.__get_worksheet(wb)
			datarow_start_index = 1
			if write_header == True:
				for col_header_index in range(0, len(data.column_names)):
					cell_header_data = data.column_names[col_header_index]
					ws.cell(row = datarow_start_index, column = col_header_index + 1).value = cell_header_data
				datarow_start_index += 1
			for rowIndex in range(0, len(data)):
				rowdata = data[rowIndex]
				for colIndex in range(0, len(rowdata)):
					celldata = rowdata[colIndex]
					ws.cell(row = rowIndex + datarow_start_index, column = colIndex + 1).value = celldata
			wb.save(self.filepath)
